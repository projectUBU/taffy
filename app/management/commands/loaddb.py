from django.core.management.base import BaseCommand, CommandError
from app.models import *

from django.contrib.auth.models import  User

from openpyxl import load_workbook


class Command(BaseCommand):
    help = 'Closes the specified poll for voting'

    def add_arguments(self, parser):
        #parser.add_argument('poll_ids', nargs='+', type=int)
        pass

    def load(self, wb, sheet_name, column_names):
        print(f'กำลัง load ... {sheet_name}')
        ws = wb[sheet_name]
        count = int(ws['A1'].value)

        print(f'count = {count}')
        data = []
        for i in range(count):  # 0,1,2,3
            # print(f'i = {i}')
            sheet_values = [
                ws[f'{chr(65+j)}{3+i}'].value for j in range(len(column_names))
            ]
            data.append(
                dict((k, v) for k, v in zip(column_names, sheet_values)))

        return data

    def handle(self, *args, **options):

        
        filename = "xlsx/loaddata.xlsx"
        wb = load_workbook(filename, data_only=True)



        for b in self.load(wb, 'BloodType', ['id', 'bloodtype']):
            print("data  =  ",b)
            q = BloodType(**b)
            q.save()
        print("seve...Blood")

        for d in self.load(wb, 'DaysOfWeek', ['id', 'daysofweek']):
            print("DaysofWeek =  ", d)
            # (**w).save()
            q = DaysOfWeek(**d)
            q.save()
            print("____________________________")

        for n in self.load(wb, 'NakSus', ['id', 'naksus']):
            print("data =  ", n)
            q = NakSus(**n)
            q.save()
            print("____________________________")
        # print("seve...NakSus")

        for d in self.load(wb, 'RaSi', ['id', 'rasi']):
            print("RaSi =  ", d)
            # (**w).save()
            q = RaSi(**d)
            q.save()
            print("____________________________")


            
        