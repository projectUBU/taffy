from django.http import request
from app.signals import save_profile
from django.core.management.base import BaseCommand, CommandError
from app.models import *
# from django.contrib.auth.hashers import make_password, HASHERS
from django.contrib.auth.models import User

from openpyxl import load_workbook
from django.contrib.auth.hashers import make_password


class Command(BaseCommand):
    help = 'Closes the specified poll for voting'

    def add_arguments(self, parser):
        #parser.add_argument('poll_ids', nargs='+', type=int)
        pass

    def load(self, wb, sheet_name, column_names):
        print(f'กำลัง load ... {sheet_name}')
        ws = wb[sheet_name]
        count = int(ws['A1'].value)

        print(f'count = {count}')
        data = []
        for i in range(count):  # 0,1,2,3
            # print(f'i = {i}')
            sheet_values = [
                ws[f'{chr(65+j)}{3+i}'].value for j in range(len(column_names))
            ]
            data.append(
                dict((k, v) for k, v in zip(column_names, sheet_values)))

        return data

    def handle(self, *args, **options):

        filename = "xlsx/loaddata.xlsx"
        wb = load_workbook(filename, data_only=True)

        for b in self.load(wb, 'Member', ['id', 'first_name', 'last_name', 'username', 'email', 'password', 'birthday', 'bloodtype', 'gender', 'testes']):
            print("Member  =  ", type(b))
            b['password'] = make_password(b['password'])
            bloodType = BloodType.objects.get(pk=b['bloodtype'])
            b.pop('bloodtype', None)
            q = Member(**b)

            q.bloodtype = bloodType
            q.save()
        print("seve...User")
        print("____________________________")

        # for d in self.load(wb, 'Gender', ['id', 'gender']):
        #     print("Gender =  ", d)
        #     # (**w).save()
        #     q = Gender(**d)
        #     q.save()
        #     print("____________________________")
        # for d in self.load(wb, 'Testes', ['id', 'testes']):
        #     print("Testes =  ", d)
        #     # (**w).save()
        #     q = Testes(**d)
        #     q.save()
        #     print("____________________________")

        # for d in self.load(wb, 'Personality', ['id', 'personality']):
        #     print("Personality =  ", d)
        #     # (**w).save()
        #     q = Personality(**d)
        #     q.save()

        # for b in self.load(wb, 'User', ['id', 'first_name', 'last_name', 'username', 'email','password',]):
        #     print("User  =  ",b)
        #     q = User(**b)
        #     q.save()
        # print("seve...User")
        # print("____________________________")

        # user = User.objects.get(pk=d['user'])
        # d.pop('user', None)
        # daysofweek = DaysOfWeek.objects.get(pk=d['daysofweek'])
        # d.pop('daysofweek',None)

        # rasi = RaSi.objects.get(pk=d['rasi'])
        # d.pop('rasi', None)

        # bloodType = BloodType.objects.get(pk=d['bloodtype'])
        # d.pop('bloodtype', None)

        # nakSus = NakSus.objects.get(pk=d['naksus'])
        # d.pop('naksus', None)

        # genDer = Gender.objects.get(pk=d['gender'])
        # d.pop('gender', None)

        # testes = Testes.objects.get(pk=d['testes'])
        # d.pop('testes', None)

        # personality = Personality.objects.get(pk=d['personality'])
        # print(f'__________count = {personality}_______________')
        # d.pop('personality', None)

        # q = Profile(**d)

        # q.user=user

        # q.rasi = rasi
        # q.daysofweek = daysofweek
        # q.bloodtype = bloodType
        # q.naksus = nakSus
        # q.gender = genDer
        # q.testes = testes

        # q.personality = personality

        # q.save()

        # for d in self.load(wb, 'Handler', ['id', 'rejected', 'reviewe_value', 'rejected']):
        #     print("Handler =  ", d)
        #     # (**w).save()
        #     q = Handler(**d)
        #     q.save()

        # for d in self.load(wb, 'Conversation', [
        #         'id', 'user', 'message', 'rejected',
        # ]):
        #     print("Conversation =  ", d)
        #     # (**w).save()

        #     user = User.objects.get(pk=d['user'])
        #     d.pop('user', None)

        #     rejected = Handler.objects.get(pk=d['rejected'])
        #     d.pop('user', None)

        #     q = Conversation(**d)
        #     q.user = user
        #     q.rejected = rejected
        #     q.save()

        # for d in self.load(wb, 'Goldmember', [
        #        'id' ,'goldmember' ,'conversation'

        # ]):
        #     print("Goldmember =  ", d)
        #     # (**w).save()
        #     goldmember = User.objects.get(pk=d['goldmember'])
        #     d.pop('goldmember', None)
        #     conversation = Conversation.objects.get(pk=d['conversation'])
        #     d.pop('conversation', None)

        #     q = Goldmember(**d)

        #     q.goldmember = goldmember
        #     q.conversation = conversation
        #     q.save()
